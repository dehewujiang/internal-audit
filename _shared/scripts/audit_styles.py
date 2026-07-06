# -*- coding: utf-8 -*-
"""
Anthropic 审计品牌样式
======================
公司品牌统一的 Excel 样式常量。

用法:
from _shared.scripts.audit_styles import AuditStyles, COLORS, FONTS

# 直接使用常量
header_font = FONTS['header']
header_fill = PatternFill(start_color=COLORS['header_blue'], ...)
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class COLORS:
    """Anthropic 品牌色"""
    # 品牌主色
    HEADER_BLUE = "2F5496"
    HEADER_WHITE = "FFFFFF"

    # 辅助色
    ROW_ALT = "D6E4F0"
    ROW_WHITE = "FFFFFF"
    RED_FLAG_BG = "C00000"
    GRAY_TEXT = "808080"
    YELLOW_WARNING = "FFE699"

    # 风险等级色
    HIGH_RISK = "C00000"
    MEDIUM_RISK = "FFC000"
    LOW_RISK = "92D050"


class FONTS:
    """字体样式"""
    header = Font(bold=True, color=COLORS.HEADER_WHITE, size=11)
    normal = Font(size=10)
    small = Font(size=9)
    gray = Font(size=10, italic=True, color=COLORS.GRAY_TEXT)
    red_flag = Font(bold=True, color=COLORS.HEADER_WHITE, size=11)
    warning = Font(bold=True, color=COLORS.HIGH_RISK, size=10)


class FILLS:
    """填充样式"""
    header = PatternFill(start_color=COLORS.HEADER_BLUE, end_color=COLORS.HEADER_BLUE, fill_type="solid")
    row_alt = PatternFill(start_color=COLORS.ROW_ALT, end_color=COLORS.ROW_ALT, fill_type="solid")
    row_white = PatternFill(start_color=COLORS.ROW_WHITE, end_color=COLORS.ROW_WHITE, fill_type="solid")
    red_flag = PatternFill(start_color=COLORS.RED_FLAG_BG, end_color=COLORS.RED_FLAG_BG, fill_type="solid")
    warning = PatternFill(start_color=COLORS.YELLOW_WARNING, end_color=COLORS.YELLOW_WARNING, fill_type="solid")


class ALIGNMENTS:
    """对齐方式"""
    header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)


class BORDERS:
    """边框样式"""
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    medium = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )


class AuditStyles:
    """
    审计样式统一入口

    用法:
    styles = AuditStyles()
    cell.font = styles.fonts.header
    cell.fill = styles.fills.header
    """

    def __init__(self):
        self.colors = COLORS
        self.fonts = FONTS
        self.fills = FILLS
        self.alignments = ALIGNMENTS
        self.borders = BORDERS

    def apply_header_style(self, cell):
        """应用表头样式"""
        cell.font = self.fonts.header
        cell.fill = self.fills.header
        cell.alignment = self.alignments.header
        cell.border = self.borders.thin

    def apply_cell_style(self, cell, is_alt_row=False):
        """应用单元格样式"""
        cell.font = self.fonts.normal
        cell.alignment = self.alignments.cell
        cell.border = self.borders.thin
        if is_alt_row:
            cell.fill = self.fills.row_alt
        else:
            cell.fill = self.fills.row_white

    def apply_warning_style(self, cell):
        """应用警示样式（用于高风险标记）"""
        cell.fill = self.fills.warning
        cell.font = self.fonts.warning


# 预计算行高
DEFAULT_ROW_HEIGHT = 25
EXPANDED_ROW_HEIGHT = 60  # 用于问题/描述类内容
COMPACT_ROW_HEIGHT = 20   # 用于序号/状态类内容
