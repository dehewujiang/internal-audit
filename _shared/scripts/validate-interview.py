#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
validate-interview.py — 审计访谈问卷 Excel 硬校验脚本

对 interview-designer 生成的访谈 Excel 文件执行确定性校验。

[INPUT]:  访谈问卷 Excel 文件路径 (.xlsx)
[OUTPUT]: JSON 格式校验报告 {"status": "pass|fail", "checks": [...], "overall": "pass|fail"}
          退出码: 0 (非--strict 总是 0), 1 (--strict 且有失败项)
[POS]:    _shared/scripts 的访谈问卷校验工具，被 audit-interview-designer/SKILL.md 引用
[PROTOCOL]: 变更时更新此头部, 然后检查同级 CLAUDE.md
"""

import json
import sys
import argparse
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl 未安装。执行: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ── 期望表头 ──────────────────────────────────────────────

SHEET1_HEADERS = ["模块", "序号", "问题", "追问提示", "制度依据", "访谈记录", "证据索引", "风险标记"]
SHEET2_HEADERS = ["序号", "资料名称", "格式", "责任部门", "是否获取", "备注"]

# 封闭式提问模式（检查 6 的逆向——这些问题算"不够开放"）
CLOSED_QUESTION_PATTERNS = [
    r'^是否',
    r'^有没有',
    r'^是不是',
]


# ── 校验函数 ─────────────────────────────────────────────

def check_file_opens(path):
    """[01] Excel 文件存在且能用 openpyxl 打开"""
    p = Path(path)
    if not p.exists():
        return False, f"文件不存在: {path}"
    if not p.suffix.lower() in (".xlsx", ".xlsm"):
        return False, f"非 Excel 文件扩展名: {p.suffix}"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
        return True, None
    except Exception as e:
        return False, f"无法打开 Excel 文件: {e}"


def _normalize_header(value):
    """去除空白后的表头文本，空单元格返回空字符串"""
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "")


def check_sheet1_headers(ws):
    """[02] Sheet 1 表头必须匹配预期列名"""
    if ws.max_column < len(SHEET1_HEADERS):
        return False, (
            f"Sheet1 列数不足（期望 ≥{len(SHEET1_HEADERS)}，实际 {ws.max_column}）"
        )
    actual = []
    for col_idx in range(1, len(SHEET1_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        actual.append(_normalize_header(cell.value))
    if actual != SHEET1_HEADERS:
        return False, f"Sheet1 表头不匹配: 期望 {SHEET1_HEADERS}, 实际 {actual}"
    return True, None


def check_sheet2_headers(ws):
    """[03] Sheet 2 (DRL) 表头必须匹配预期列名"""
    if ws.max_column < len(SHEET2_HEADERS):
        return False, (
            f"Sheet2 列数不足（期望 ≥{len(SHEET2_HEADERS)}，实际 {ws.max_column}）"
        )
    actual = []
    for col_idx in range(1, len(SHEET2_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        actual.append(_normalize_header(cell.value))
    if actual != SHEET2_HEADERS:
        return False, f"Sheet2 表头不匹配: 期望 {SHEET2_HEADERS}, 实际 {actual}"
    return True, None


def check_sheet3_exists(sheet_names):
    """[04] Sheet 3 必须存在（不检查表头格式）"""
    # openpyxl 索引从 0 开始 → Sheet 3 = 索引 2
    if len(sheet_names) < 3:
        return False, f"Sheet3 不存在（工作表总数: {len(sheet_names)}）"
    return True, None


def check_sheet4_exists(sheet_names):
    """[05] Sheet 4 可选，存在即通过（不检查表头格式）"""
    if len(sheet_names) < 4:
        return True, "Sheet4 不存在（可选，非错误）"
    return True, None


def check_question_count(ws):
    """[06] 问题列 (Sheet1 C列) 至少有 5 个非空行"""
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value is not None and str(cell.value).strip():
            count += 1
    if count < 5:
        return False, f"问题列非空行不足: {count}（需要 ≥5）"
    return True, f"问题列共 {count} 个非空行"


def check_open_question_ratio(ws):
    """[07] 非封闭式问题（非 是否/有没有/是不是）占比 ≥ 70%"""
    total = 0
    closed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        text = str(cell.value).strip() if cell.value is not None else ""
        if not text:
            continue
        total += 1
        if any(re.match(p, text) for p in CLOSED_QUESTION_PATTERNS):
            closed += 1
    if total == 0:
        return False, "问题列无有效内容，无法计算开放问题比例"
    open_count = total - closed
    ratio = open_count / total * 100
    if ratio < 70:
        return False, (
            f"开放问题比例不足: {ratio:.1f}% "
            f"（开放 {open_count} / 总计 {total}，需要 ≥70%）"
        )
    return True, f"开放问题比例 {ratio:.1f}% ({open_count}/{total})"


def check_policy_reference_ratio(ws):
    """[08] 制度依据列 (Sheet1 E列) 非空比例 ≥ 50%"""
    total = 0
    filled = 0
    for row_idx in range(2, ws.max_row + 1):
        # 只统计有问题的行
        q_cell = ws.cell(row=row_idx, column=3)
        q_text = str(q_cell.value).strip() if q_cell.value is not None else ""
        if not q_text:
            continue
        total += 1
        e_cell = ws.cell(row=row_idx, column=5)
        e_text = str(e_cell.value).strip() if e_cell.value is not None else ""
        if e_text:
            filled += 1
    if total == 0:
        return False, "无有效问题行，无法计算制度依据覆盖率"
    ratio = filled / total * 100
    if ratio < 50:
        return False, (
            f"制度依据覆盖率不足: {ratio:.1f}% "
            f"（已填写 {filled} / 有问题的行 {total}，需要 ≥50%）"
        )
    return True, f"制度依据覆盖率 {ratio:.1f}% ({filled}/{total})"


def check_drl_entry_count(ws):
    """[09] 资料需求清单 (Sheet2) 至少 3 条记录"""
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        # 序数列 (A列) 或 资料名称列 (B列) 有任意非空即算一条
        a_cell = ws.cell(row=row_idx, column=1)
        b_cell = ws.cell(row=row_idx, column=2)
        a_ok = a_cell.value is not None and str(a_cell.value).strip()
        b_ok = b_cell.value is not None and str(b_cell.value).strip()
        if a_ok or b_ok:
            count += 1
    if count < 3:
        return False, f"资料需求清单条目不足: {count}（需要 ≥3）"
    return True, f"资料需求清单共 {count} 条"


# ── 主校验流程 ───────────────────────────────────────────

def validate_interview(path):
    """执行全部 10 项校验，返回 (checks_list, overall_pass)"""
    checks = []

    # [01] 文件可打开
    passed, detail = check_file_opens(path)
    checks.append({"name": "file_opens", "result": "pass" if passed else "fail", "detail": detail or "文件可正常打开"})

    if not passed:
        # 文件打不开，后续检查无法进行
        checks.extend([
            {"name": "sheet1_headers", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "sheet2_headers", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "sheet3_exists", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "sheet4_exists", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "question_count", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "open_question_ratio", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "policy_reference_ratio", "result": "fail", "detail": "前置检查失败：文件无法打开"},
            {"name": "drl_entry_count", "result": "fail", "detail": "前置检查失败：文件无法打开"},
        ])
        overall = False
        return checks, overall

    # 加载工作簿
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws1 = wb[sheet_names[0]] if len(sheet_names) >= 1 else None
    ws2 = wb[sheet_names[1]] if len(sheet_names) >= 2 else None

    # [02] Sheet1 表头
    if ws1 is None:
        passed, detail = False, "Sheet1 不存在"
    else:
        passed, detail = check_sheet1_headers(ws1)
    checks.append({"name": "sheet1_headers", "result": "pass" if passed else "fail",
                   "detail": detail or "Sheet1 表头匹配"})

    # [03] Sheet2 表头
    if ws2 is None:
        passed, detail = False, "Sheet2 不存在"
    else:
        passed, detail = check_sheet2_headers(ws2)
    checks.append({"name": "sheet2_headers", "result": "pass" if passed else "fail",
                   "detail": detail or "Sheet2 表头匹配"})

    # [04] Sheet3 存在
    passed, detail = check_sheet3_exists(sheet_names)
    checks.append({"name": "sheet3_exists", "result": "pass" if passed else "fail",
                   "detail": detail or "Sheet3 存在"})

    # [05] Sheet4 可选
    passed, detail = check_sheet4_exists(sheet_names)
    checks.append({"name": "sheet4_exists", "result": "pass" if passed else "fail",
                   "detail": detail or "Sheet4 存在（可选）"})

    # [06] 问题数量 ≥ 5
    if ws1 is None:
        passed, detail = False, "Sheet1 不存在"
    else:
        passed, detail = check_question_count(ws1)
    checks.append({"name": "question_count", "result": "pass" if passed else "fail",
                   "detail": detail or "问题数量达标"})

    # [07] 开放问题比例
    if ws1 is None:
        passed, detail = False, "Sheet1 不存在"
    else:
        passed, detail = check_open_question_ratio(ws1)
    checks.append({"name": "open_question_ratio", "result": "pass" if passed else "fail",
                   "detail": detail or "开放问题比例达标"})

    # [08] 制度依据覆盖率
    if ws1 is None:
        passed, detail = False, "Sheet1 不存在"
    else:
        passed, detail = check_policy_reference_ratio(ws1)
    checks.append({"name": "policy_reference_ratio", "result": "pass" if passed else "fail",
                   "detail": detail or "制度依据覆盖率达标"})

    # [09] DRL 条目数
    if ws2 is None:
        passed, detail = False, "Sheet2 不存在"
    else:
        passed, detail = check_drl_entry_count(ws2)
    checks.append({"name": "drl_entry_count", "result": "pass" if passed else "fail",
                   "detail": detail or "资料需求清单条目数达标"})

    wb.close()

    # [10] 汇总判定
    overall = all(c["result"] == "pass" for c in checks)

    return checks, overall


# ── CLI ─────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="审计访谈问卷 Excel 硬校验工具"
    )
    parser.add_argument(
        "file",
        help="访谈问卷 Excel 文件路径 (.xlsx / .xlsm)"
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="严格模式：任一检查失败 → exit(1)"
    )
    args = parser.parse_args()

    checks, overall = validate_interview(args.file)

    report = {
        "status": "pass" if overall else "fail",
        "checks": checks,
        "overall": "pass" if overall else "fail",
    }

    # 输出 JSON（无论 --strict，总是打印）
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if args.strict and not overall:
        sys.exit(1)

    sys.exit(0)


# ── 入口 ─────────────────────────────────────────────────

if __name__ == "__main__":
    main()
