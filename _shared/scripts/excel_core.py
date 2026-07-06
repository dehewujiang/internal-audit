# -*- coding: utf-8 -*-
"""
Excel 核心操作
==============
通用的 Excel 文件创建和操作工具，不包含业务逻辑。

用法:
import sys
from pathlib import Path
_SHARED = Path(__file__).resolve().parent
sys.path.insert(0, str(_SHARED))

from excel_core import ExcelCore

core = ExcelCore("output.xlsx")
core.add_sheet("Sheet1", headers=["A", "B"], rows=[[1, 2]])
core.save()
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Union
import os
import re

from audit_styles import AuditStyles, COLORS, FONTS, FILLS, ALIGNMENTS, BORDERS


def convert_br_to_newline(text: str) -> str:
    """将 HTML <br> 标签转换为换行符"""
    if not isinstance(text, str):
        return text
    return re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)


class ExcelCore:
    """
    Excel 核心操作类

    职责：纯粹的 Excel 文件操作，不包含任何审计业务逻辑
    """

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        self._sheet_count = 0
        self.styles = AuditStyles()

    def add_worksheet(
        self,
        title: str,
        headers: List[str],
        rows: List[List],
        col_widths: Optional[List[int]] = None,
        row_heights: Optional[Dict[int, int]] = None,
        freeze_header: bool = True,
        alt_row_colors: bool = True,
        default_height: int = 25,
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        添加工作表

        Args:
            title: Sheet 标题
            headers: 表头列表
            rows: 数据行（2维列表）
            col_widths: 列宽列表
            row_heights: 行高字典 {行号: 高度}
            freeze_header: 是否冻结首行
            alt_row_colors: 是否交替行颜色
            default_height: 默认行高
        """
        # 处理 Sheet 名重复
        self._sheet_count += 1
        if self._sheet_count == 1:
            ws = self.wb.active
            ws.title = title
        else:
            # 处理重复名称
            base_title = title
            existing = [s.title for s in self.wb.worksheets]
            if title in existing:
                counter = 1
                new_title = f"{title}_{counter}"
                while new_title in existing:
                    counter += 1
                    new_title = f"{title}_{counter}"
                title = new_title
            ws = self.wb.create_sheet(title)

        # 设置列宽
        if col_widths:
            for i, width in enumerate(col_widths, 1):
                letter = get_column_letter(i)
                ws.column_dimensions[letter].width = width
        else:
            # 自动计算列宽
            for i, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row in rows:
                    if i <= len(row):
                        content = str(row[i - 1]).split('\n')[0]
                        max_len = max(max_len, min(len(content), 50))
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.styles.apply_header_style(cell)

        # 冻结首行
        if freeze_header:
            ws.freeze_panes = "A2"

        # 写入数据
        for row_idx, row_data in enumerate(rows, 2):
            is_alt = alt_row_colors and (row_idx % 2 == 0)
            for col_idx, value in enumerate(row_data, 1):
                # 处理换行符
                value = convert_br_to_newline(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self.styles.apply_cell_style(cell, is_alt_row=is_alt)

            # 设置行高
            if row_heights and row_idx in row_heights:
                ws.row_dimensions[row_idx].height = row_heights[row_idx]
            else:
                ws.row_dimensions[row_idx].height = default_height

        return ws

    def save(self) -> str:
        """保存文件，返回路径"""
        os.makedirs(os.path.dirname(self.output_path) or ".", exist_ok=True)
        self.wb.save(self.output_path)
        return self.output_path

    def get_active_sheet(self):
        """获取当前活动 Sheet"""
        return self.wb.active

    def get_sheet_names(self) -> List[str]:
        """获取所有 Sheet 名称"""
        return [s.title for s in self.wb.worksheets]

    def close(self):
        """关闭工作簿（释放内存）"""
        if self.wb:
            self.wb.close()


class SheetBuilder:
    """链式 Sheet 构建器（简化复杂 Sheet 的创建）"""

    def __init__(self, core: ExcelCore, title: str):
        self.core = core
        self.title = title
        self.headers: List[str] = []
        self.rows: List[List] = []
        self.col_widths: List[int] = []
        self.row_heights: Dict[int, int] = {}

    def set_headers(self, headers: List[str], widths: Optional[List[int]] = None):
        """设置表头和列宽"""
        self.headers = headers
        self.col_widths = widths or []
        return self

    def add_row(self, row: List, height: int = 25):
        """添加一行数据"""
        self.rows.append(row)
        self.row_heights[len(self.rows) + 1] = height  # +1 因为表头在第1行
        return self

    def add_rows(self, rows: List[List], height: int = 25):
        """批量添加行"""
        for row in rows:
            self.add_row(row, height)
        return self

    def build(self, **kwargs) -> openpyxl.worksheet.worksheet.Worksheet:
        """构建并返回 Sheet"""
        return self.core.add_worksheet(
            title=self.title,
            headers=self.headers,
            rows=self.rows,
            col_widths=self.col_widths,
            row_heights=self.row_heights,
            **kwargs
        )


def create_excel(output_path: str, sheets: List[Dict]) -> str:
    """
    快速创建多 Sheet Excel

    Args:
        output_path: 输出路径
        sheets: Sheet 配置列表
            [
                {
                    "title": "Sheet1",
                    "headers": ["A", "B"],
                    "rows": [[1, 2]],
                    "col_widths": [15, 15]
                }
            ]
    """
    core = ExcelCore(output_path)
    for sheet_config in sheets:
        core.add_worksheet(
            title=sheet_config["title"],
            headers=sheet_config["headers"],
            rows=sheet_config["rows"],
            col_widths=sheet_config.get("col_widths"),
        )
    return core.save()
